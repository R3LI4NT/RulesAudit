import pandas as pd
from datetime import datetime
import tkinter as tk
from tkinter import filedialog
import os
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
import chardet

# Colores
RED = '\033[1;31m'
GREEN = '\033[1;32m'
YELLOW = '\033[1;33m'
END = '\033[0m'

# ===============================
# FUNCIÓN PARA DETECTAR CODIFICACIÓN
# ===============================

def detectar_codificacion(archivo):
    with open(archivo, 'rb') as f:
        raw_data = f.read(10000)  
        resultado = chardet.detect(raw_data)
        return resultado['encoding']


def leer_csv_con_codificacion(ruta_csv):
    # Lista extendida de codificaciones comunes
    codificaciones = [
        'utf-8',
        'utf-8-sig',
        'utf-16',
        'utf-16-le',
        'utf-16-be',
        'utf-32',
        'utf-32-le',
        'utf-32-be',
        'latin-1',
        'iso-8859-1',
        'iso-8859-15',
        'cp1252',
        'cp850',
        'cp437',
        'cp1250',
        'cp1254',
        'iso-8859-2',
        'iso-8859-9',
        'mac_roman',
        'mac_latin2'
    ]
    
    try:
        cod_detectada = detectar_codificacion(ruta_csv)
        if cod_detectada:
            codificaciones.insert(0, cod_detectada)
    except:
        pass
    
    for encoding in codificaciones:
        try:
            df = pd.read_csv(ruta_csv, skipinitialspace=True, encoding=encoding)
            return df, encoding
        except UnicodeDecodeError:
            continue
        except Exception as e:
            continue
    
    try:
        df = pd.read_csv(ruta_csv, skipinitialspace=True, encoding='utf-8', errors='ignore')
        return df, 'utf-8 (ignorando errores)'
    except Exception as e:
        return None, None


# ===============================
# SELECTORES 
# ===============================

def seleccionar_archivo_csv():
    raiz = tk.Tk()
    raiz.withdraw()
    raiz.lift()
    raiz.focus_force()

    print("\nPresiona ENTER para seleccionar el archivo CSV...")
    input()

    ruta_csv = filedialog.askopenfilename(
        title="Selecciona el archivo CSV",
        filetypes=[("Archivos CSV", "*.csv")]
    )

    raiz.destroy()
    return ruta_csv


def seleccionar_carpeta_csv():
    raiz = tk.Tk()
    raiz.withdraw()
    raiz.lift()
    raiz.focus_force()

    print("\nPresiona ENTER para seleccionar la carpeta donde están los CSV...")
    input()

    carpeta = filedialog.askdirectory(
        title="Selecciona la carpeta que contiene los CSV"
    )

    raiz.destroy()
    return carpeta


def seleccionar_carpeta_destino():
    raiz = tk.Tk()
    raiz.withdraw()
    raiz.lift()
    raiz.focus_force()

    print("\nPresiona ENTER para seleccionar la carpeta de destino...")
    input()

    carpeta = filedialog.askdirectory(
        title="Selecciona la carpeta donde guardar el Excel"
    )

    raiz.destroy()
    return carpeta


# ===============================
# FORMATO EXCEL 
# ===============================

def aplicar_formato_excel(libro_excel, nombre_hoja):

    hoja = libro_excel[nombre_hoja]

    relleno_header = PatternFill(start_color="006400", end_color="006400", fill_type="solid")  # Verde fuerte
    relleno_verde_suave = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    relleno_blanco = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    for celda in hoja[1]:
        celda.fill = relleno_header
        celda.font = Font(color="FFFFFF", bold=True)
        celda.alignment = Alignment(horizontal="center", vertical="center")

    for fila in range(2, hoja.max_row + 1):
        relleno = relleno_verde_suave if fila % 2 == 0 else relleno_blanco
        for col in range(1, hoja.max_column + 1):
            hoja.cell(row=fila, column=col).fill = relleno

    for columna in hoja.columns:
        max_len = 0
        letra = get_column_letter(columna[0].column)
        for celda in columna:
            if celda.value:
                max_len = max(max_len, len(str(celda.value)))
        hoja.column_dimensions[letra].width = max_len + 2

    hoja.freeze_panes = "A2"
    hoja.auto_filter.ref = hoja.dimensions


# ===============================
# PROCESAR UN SOLO CSV 
# ===============================

def mapear_csv_a_excel(ruta_csv, carpeta_destino):

    print(f"\n{GREEN}[+]{END} Procesando: {ruta_csv}")

    df, encoding_usado = leer_csv_con_codificacion(ruta_csv)
    
    if df is None:
        print(f"{YELLOW}[!]{END} No se pudo leer el archivo con ninguna codificación")
        return False

    print(f"{GREEN}[+]{END} Codificación utilizada: {encoding_usado}")

    ruta_salida = os.path.join(
        carpeta_destino,
        f"{os.path.splitext(os.path.basename(ruta_csv))[0]}_FORMATEADO_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    )

    with pd.ExcelWriter(ruta_salida, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Resultado", index=False)
        aplicar_formato_excel(writer.book, "Resultado")

    print(f"{GREEN}[+]{END} Excel generado en: {ruta_salida}")
    return True


# ===============================
# PROCESAR LOTE COMPLETO
# ===============================

def mapear_lote_a_excel(carpeta_csv, carpeta_destino):

    print(f"\n{GREEN}[+]{END} Buscando archivos CSV en la carpeta...")
    print("-" * 50)

    rutas_csv = [
        os.path.join(carpeta_csv, f)
        for f in os.listdir(carpeta_csv)
        if f.lower().endswith(".csv")
    ]

    if not rutas_csv:
        print(f"{YELLOW}[!]{END} No se encontraron archivos CSV")
        return False

    print(f"{GREEN}[+]{END} Se encontraron {len(rutas_csv)} archivos")

    todas_las_filas = []
    archivos_procesados = 0
    archivos_con_error = []

    for ruta in rutas_csv:
        print(f"{GREEN}[+]{END} Cargando: {os.path.basename(ruta)}")

        df, encoding_usado = leer_csv_con_codificacion(ruta)
        
        if df is not None:
            print(f"    → Codificación: {encoding_usado}")
            for _, fila in df.iterrows():
                todas_las_filas.append(fila.tolist())
            archivos_procesados += 1
        else:
            print(f"{RED}[X]{END} Error leyendo {os.path.basename(ruta)}")
            archivos_con_error.append(os.path.basename(ruta))

    if archivos_con_error:
        print(f"\n{RED}[X]{END} Archivos con error: {len(archivos_con_error)}")
        for archivo in archivos_con_error:
            print(f"    - {archivo}")

    if not todas_las_filas:
        print(f"{YELLOW}[!]{END} No hay datos para procesar")
        return False

    max_columnas = max(len(f) for f in todas_las_filas)

    columnas_base = [
        'Policy', 'Source', 'Destination', 'Schedule', 'Service',
        'Action', 'IP Pool', 'NAT', 'Type', 'No.', 'Type', 'Hits',
        'First Hits', 'Last Hits', 'Name', 'VPN', 'Services & Applications',
        'Track' 'Install On', 'Uid' 'Security Profiles', 'Log', 'Bytes'
    ]

    if max_columnas > len(columnas_base):
        for i in range(len(columnas_base), max_columnas):
            columnas_base.append(f"Extra_{i+1}")

    filas_completadas = [
        f + [''] * (max_columnas - len(f))
        for f in todas_las_filas
    ]

    df_final = pd.DataFrame(
        filas_completadas,
        columns=columnas_base[:max_columnas]
    )

    ruta_salida = os.path.join(
        carpeta_destino,
        f"Lote_Unificado_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    )

    with pd.ExcelWriter(ruta_salida, engine="openpyxl") as writer:
        df_final.to_excel(writer, sheet_name="Unificado", index=False)
        aplicar_formato_excel(writer.book, "Unificado")

    print(f"\n{GREEN}[+]{END} Excel unificado generado en: {ruta_salida}")
    print(f"{GREEN}[+]{END} Archivos procesados exitosamente: {archivos_procesados} de {len(rutas_csv)}")
    return True


# ===============================
# MAIN
# ===============================

def main():

    print("=" * 60)
    print("[+]  MAPEADOR DE REGLAS DE FIREWALL [+]")
    print("=" * 60)

    print("\nSelecciona el modo de trabajo:")
    print("1 - Procesar un solo CSV")
    print("2 - Procesar carpeta completa (unificado)")

    opcion = input("\nOpción (1 o 2): ")

    if opcion == "1":

        ruta_csv = seleccionar_archivo_csv()
        if not ruta_csv:
            print(f"{YELLOW}[!]{END} Archivo inválido")
            return

        carpeta_destino = seleccionar_carpeta_destino()
        if not carpeta_destino:
            print(f"{YELLOW}[!]{END} Carpeta inválida")
            return

        mapear_csv_a_excel(ruta_csv, carpeta_destino)

    elif opcion == "2":

        carpeta_csv = seleccionar_carpeta_csv()
        if not carpeta_csv:
            print(f"{YELLOW}[!]{END} Carpeta inválida")
            return

        carpeta_destino = seleccionar_carpeta_destino()
        if not carpeta_destino:
            print(f"{YELLOW}[!]{END} Carpeta inválida")
            return

        mapear_lote_a_excel(carpeta_csv, carpeta_destino)

    else:
        print(f"{YELLOW}[!]{END} Opción inválida")


if __name__ == "__main__":
    main()