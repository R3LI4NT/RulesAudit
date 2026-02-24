#Módulo para generación de reportes Excel

#Colores
RED = '\033[1;31m'
GREEN = '\033[1;32m'
YELLOW = '\033[1;33m'
END = '\033[0m'

import pandas as pd
from datetime import datetime
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter


def generar_reporte_excel(vulnerabilidades, ruta_salida):
    if not vulnerabilidades:
        print(f"\n{GREEN}[+]{END} No se encontraron vulnerabilidades")
        return
    
    df_vuln = pd.DataFrame(vulnerabilidades)
    column_order = ['Severidad', 'Tipo', 'Sección', 'Regla', 'Descripción', 'Recomendación']
    columnas_existentes = [col for col in column_order if col in df_vuln.columns]
    df_vuln = df_vuln[columnas_existentes]
    
    # Estadísticas por severidad
    stats = df_vuln['Severidad'].value_counts()
    
    with pd.ExcelWriter(ruta_salida, engine='openpyxl') as writer:
        df_vuln.to_excel(writer, sheet_name='Vulnerabilidades', index=False)
        
        # Hoja de resumen
        resumen = pd.DataFrame({
            'Métrica': [
                'Total Vulnerabilidades',
                '🔴🔴 CRÍTICA',
                '🔴 ALTA', 
                '🟠 MEDIA',
                '🟡 BAJA',
                '🔵 INFORMATIVA',
                'Fecha Análisis',
                'Archivo Analizado'
            ],
            'Valor': [
                len(vulnerabilidades),
                stats.get('CRÍTICA', 0),
                stats.get('ALTA', 0),
                stats.get('MEDIA', 0),
                stats.get('BAJA', 0),
                stats.get('INFORMATIVA', 0),
                datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                ruta_salida.split('/')[-1].replace('Reporte_Vulnerabilidades_', '').replace('.xlsx', '')
            ]
        })
        resumen.to_excel(writer, sheet_name='Resumen', index=False)
        
        # Hoja de top vulnerabilidades
        if 'Descripción' in df_vuln.columns:
            top_vuln = df_vuln['Descripción'].value_counts().head(10).reset_index()
            top_vuln.columns = ['Descripción', 'Frecuencia']
            top_vuln.to_excel(writer, sheet_name='Top_Vulnerabilidades', index=False)
        
        # Aplicar formato
        aplicar_formato_reporte(writer.book, 'Vulnerabilidades')
        aplicar_formato_reporte(writer.book, 'Resumen')
        if 'Top_Vulnerabilidades' in writer.book.sheetnames:
            aplicar_formato_reporte(writer.book, 'Top_Vulnerabilidades')
    
    print(f"\n{GREEN}[+]{END} Reporte Excel generado: {ruta_salida}")
    print(f"\n{GREEN}[+]{END} RESUMEN DE VULNERABILIDADES:")
    print("-" * 50)
    print(f"Total: {len(vulnerabilidades)}")
    print(f"🔴🔴 CRÍTICA: {stats.get('CRÍTICA', 0)}")
    print(f"🔴 ALTA: {stats.get('ALTA', 0)}")
    print(f"🟠 MEDIA: {stats.get('MEDIA', 0)}")
    print(f"🟡 BAJA: {stats.get('BAJA', 0)}")
    print(f"🔵 INFORMATIVA: {stats.get('INFORMATIVA', 0)}")


def aplicar_formato_reporte(libro_excel, nombre_hoja):
    
    if nombre_hoja not in libro_excel.sheetnames:
        return
        
    hoja = libro_excel[nombre_hoja]
    
    # Colores por severidad
    colores = {
        'CRÍTICA': PatternFill(start_color="8B0000", end_color="8B0000", fill_type="solid"),  # Rojo oscuro
        'ALTA': PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid"),      # Rojo
        'MEDIA': PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid"),     # Naranja
        'BAJA': PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid"),      # Verde claro
        'INFORMATIVA': PatternFill(start_color="87CEEB", end_color="87CEEB", fill_type="solid") # Celeste
    }
    
    header_fill = PatternFill(start_color="1E3C72", end_color="1E3C72", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    
    for celda in hoja[1]:
        celda.fill = header_fill
        celda.font = header_font
        celda.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Colorear por severidad 
    if nombre_hoja == 'Vulnerabilidades':
        for fila in range(2, hoja.max_row + 1):
            severidad = hoja.cell(row=fila, column=1).value
            if severidad in colores:
                for col in range(1, hoja.max_column + 1):
                    celda = hoja.cell(row=fila, column=col)
                    celda.fill = colores[severidad]
                    if severidad in ['CRÍTICA', 'ALTA', 'MEDIA']:
                        celda.font = Font(color="FFFFFF")
                    else:
                        celda.font = Font(color="000000")
    
    for columna in hoja.columns:
        max_len = 0
        letra = get_column_letter(columna[0].column)
        for celda in columna:
            if celda.value:
                max_len = max(max_len, len(str(celda.value)))
        hoja.column_dimensions[letra].width = min(max_len + 2, 60)
    
    hoja.freeze_panes = "A2"
    hoja.auto_filter.ref = hoja.dimensions