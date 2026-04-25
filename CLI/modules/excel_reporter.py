import pandas as pd
from datetime import datetime
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


COLORES_SEVERIDAD = {
    'CRÍTICA': '8B0000',
    'ALTA': 'FF0000',
    'MEDIA': 'FFA500',
    'BAJA': '90EE90',
    'INFORMATIVA': '87CEEB'
}


def generar_reporte_excel(vulnerabilidades, ruta_salida, resumen=None, score=None, compliance=None):
    if not vulnerabilidades:
        return None

    df_vuln = pd.DataFrame(vulnerabilidades)
    columnas_vuln = ['ID', 'Severidad', 'CVSS', 'Tipo', 'Categoria', 'Sección', 'Regla', 'Descripción', 'Recomendación']
    columnas_existentes = [c for c in columnas_vuln if c in df_vuln.columns]
    df_vuln_main = df_vuln[columnas_existentes].copy()

    stats = df_vuln['Severidad'].value_counts()

    filas_evidencia = []
    for v in vulnerabilidades:
        evidencia = v.get('Evidencia', {})
        fila = {
            'ID_Vuln': v.get('ID', ''),
            'Severidad': v.get('Severidad', ''),
            'Regla': v.get('Regla', ''),
            'Sección': v.get('Sección', '')
        }
        for campo, valor in evidencia.items():
            fila[f"Ev_{campo}"] = valor
        filas_evidencia.append(fila)
    df_evidencia = pd.DataFrame(filas_evidencia)

    with pd.ExcelWriter(ruta_salida, engine='openpyxl') as writer:
        df_vuln_main.to_excel(writer, sheet_name='Vulnerabilidades', index=False)
        df_evidencia.to_excel(writer, sheet_name='Evidencia_Reglas', index=False)

        resumen_filas = [
            ['Total Vulnerabilidades', len(vulnerabilidades)],
            ['CRITICA', stats.get('CRÍTICA', 0)],
            ['ALTA', stats.get('ALTA', 0)],
            ['MEDIA', stats.get('MEDIA', 0)],
            ['BAJA', stats.get('BAJA', 0)],
            ['INFORMATIVA', stats.get('INFORMATIVA', 0)],
            ['', ''],
            ['Fecha Analisis', datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
            ['Generador', 'RulesAudit v2.1']
        ]
        if score:
            resumen_filas.extend([
                ['', ''],
                ['Score de Riesgo', f"{score.get('score', 0)}/100"],
                ['Nivel', score.get('nivel', 'N/A')],
                ['Grado', score.get('grado', 'N/A')],
                ['Descripcion', score.get('descripcion', '')]
            ])
        if resumen:
            resumen_filas.extend([
                ['', ''],
                ['Formato Detectado', resumen.get('formato_detectado', 'N/A')],
                ['Total Reglas', resumen.get('total_reglas', 0)],
                ['Reglas Afectadas', resumen.get('reglas_afectadas', 0)],
                ['CVSS Promedio', resumen.get('cvss_promedio', 0)],
                ['CVSS Maximo', resumen.get('cvss_maximo', 0)]
            ])
        df_resumen = pd.DataFrame(resumen_filas, columns=['Metrica', 'Valor'])
        df_resumen.to_excel(writer, sheet_name='Resumen_Ejecutivo', index=False)

        if 'Descripción' in df_vuln.columns:
            top_vuln = df_vuln['Descripción'].value_counts().head(15).reset_index()
            top_vuln.columns = ['Descripción', 'Frecuencia']
            top_vuln.to_excel(writer, sheet_name='Top_Vulnerabilidades', index=False)

        if 'Categoria' in df_vuln.columns:
            por_cat = df_vuln['Categoria'].value_counts().reset_index()
            por_cat.columns = ['Categoria', 'Cantidad']
            por_cat.to_excel(writer, sheet_name='Por_Categoria', index=False)

        if compliance:
            filas_comp = []
            for marco in compliance:
                for ctrl in marco.get('controles_incumplidos', []):
                    filas_comp.append({
                        'Marco': marco.get('marco', ''),
                        'Descripcion_Marco': marco.get('descripcion', ''),
                        'Control_Incumplido': ctrl
                    })
            if filas_comp:
                df_comp = pd.DataFrame(filas_comp)
                df_comp.to_excel(writer, sheet_name='Compliance', index=False)

        aplicar_formato(writer.book, 'Vulnerabilidades', es_vuln=True)
        aplicar_formato(writer.book, 'Evidencia_Reglas', es_vuln=False, col_severidad=2)
        aplicar_formato(writer.book, 'Resumen_Ejecutivo')
        if 'Top_Vulnerabilidades' in writer.book.sheetnames:
            aplicar_formato(writer.book, 'Top_Vulnerabilidades')
        if 'Por_Categoria' in writer.book.sheetnames:
            aplicar_formato(writer.book, 'Por_Categoria')
        if 'Compliance' in writer.book.sheetnames:
            aplicar_formato(writer.book, 'Compliance')

    return ruta_salida


def aplicar_formato(libro, nombre_hoja, es_vuln=False, col_severidad=2):
    if nombre_hoja not in libro.sheetnames:
        return
    hoja = libro[nombre_hoja]

    header_fill = PatternFill(start_color="1E3C72", end_color="1E3C72", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    header_border = Border(
        left=Side(style='thin', color='2D333B'),
        right=Side(style='thin', color='2D333B'),
        top=Side(style='thin', color='2D333B'),
        bottom=Side(style='medium', color='00FFFF')
    )

    for celda in hoja[1]:
        celda.fill = header_fill
        celda.font = header_font
        celda.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        celda.border = header_border

    if es_vuln or nombre_hoja == 'Evidencia_Reglas':
        for fila in range(2, hoja.max_row + 1):
            severidad = hoja.cell(row=fila, column=col_severidad).value
            if severidad in COLORES_SEVERIDAD:
                color = COLORES_SEVERIDAD[severidad]
                fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                for col in range(1, hoja.max_column + 1):
                    celda = hoja.cell(row=fila, column=col)
                    celda.fill = fill
                    if severidad in ('CRÍTICA', 'ALTA', 'MEDIA'):
                        celda.font = Font(color="FFFFFF")
                    else:
                        celda.font = Font(color="000000")

    for columna in hoja.columns:
        max_len = 0
        letra = get_column_letter(columna[0].column)
        for celda in columna:
            if celda.value:
                valor_str = str(celda.value)
                max_len = max(max_len, min(len(valor_str), 80))
        hoja.column_dimensions[letra].width = min(max_len + 2, 60)

    hoja.freeze_panes = "A2"
    if hoja.max_row > 1:
        hoja.auto_filter.ref = hoja.dimensions
